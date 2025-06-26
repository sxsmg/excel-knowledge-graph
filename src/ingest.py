# ingest.py

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter
import networkx as nx
from .parser import extract_dependencies

def expand_range(start: str, end: str):
    """Given "A1","B3" returns all cells in that rectangle."""
    min_col, min_row, max_col, max_row = range_boundaries(f"{start}:{end}")
    cells = []
    for col in range(min_col, max_col+1):
        letter = get_column_letter(col)
        for row in range(min_row, max_row+1):
            cells.append(f"{letter}{row}")
    return cells

def build_nx_graph(path: str) -> nx.DiGraph:
    """
    Reads every sheet in the .xlsx, parses formulas (including ranges),
    and returns a directed graph G where edges are PRECEDENT → DEPENDENT.
    Node IDs are 'SheetName!A1'.
    """
    wb = load_workbook(path, data_only=False)
    G = nx.DiGraph()

    # 1) Create a node for every cell in every sheet
    for ws in wb.worksheets:
        sheet = ws.title
        for row in ws.iter_rows():
            for cell in row:
                addr = f"{sheet}!{cell.coordinate}"
                G.add_node(addr)

    # 2) Walk again to extract formula deps
    for ws in wb.worksheets:
        sheet = ws.title
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("="):
                    dst = f"{sheet}!{cell.coordinate}"
                    for sheet_ref, coord in extract_dependencies(val):
                        ref_sheet = sheet_ref or sheet
                        if ":" in coord:
                            start, end = coord.split(":")
                            for c in expand_range(start, end):
                                src = f"{ref_sheet}!{c}"
                                if src != dst:
                                    G.add_edge(src, dst)
                        else:
                            src = f"{ref_sheet}!{coord}"
                            if src != dst:
                                G.add_edge(src, dst)
    return G
