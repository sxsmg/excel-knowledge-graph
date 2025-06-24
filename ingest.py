# ingest.py

import networkx as nx
from openpyxl import load_workbook
from parser import extract_dependencies

def ingest_xlsx(path):
    wb = load_workbook(path, data_only=False)
    G = nx.DiGraph()

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                addr = f"{sheet.title}!{cell.coordinate}"
                G.add_node(addr, formula=cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None)

                if cell.data_type == 'f':  # has a formula
                    deps = extract_dependencies(cell.value)
                    for ref in deps:
                        ref_addr = f"{sheet.title}!{ref}"
                        G.add_edge(ref_addr, addr)

    return G

if __name__ == "__main__":
    graph = ingest_xlsx("Test Sheet 1.xlsx")
    print("Sample dependencies:")
    for u, v in list(graph.edges())[:10]:
        print(f"  {u} → {v}")
